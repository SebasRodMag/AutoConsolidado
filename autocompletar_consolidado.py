"""
Script para autocompletar y consolidar datos en un archivo Excel maestro.

Este script lee un archivo Excel llamado 'CONSOLIDADO.xlsx', busca códigos
en la Columna D de hojas específicas ('EDU', 'HOSP', 'EMPRESA'), y utiliza
esos códigos para encontrar y leer archivos Excel individuales
('[CODIGO].xlsx'). Los datos extraídos de celdas específicas en estos
archivos individuales se copian luego en las columnas correspondientes
(F, G, I, K, M) del archivo 'CONSOLIDADO.xlsx'.

Al finalizar, el script guarda el archivo modificado con un nuevo nombre,
'CONSOLIDADO_COMPLETADO.xlsx'.

NOTA: Esta versión del script carga el archivo 'CONSOLIDADO.xlsx'
directamente en memoria para su modificación. Para archivos muy grandes,
esto puede causar problemas de memoria (MemoryError). Se recomienda
utilizar una versión más reciente que maneja grandes archivos en modo
de solo lectura y genera un nuevo archivo de salida.
"""

import openpyxl
import os

# --- Constantes de Configuración ---

# Define la columna donde se espera encontrar el código en las hojas objetivo (D = 4)
COL_CODIGO = 4  # Columna D

# Mapea las columnas de destino en 'CONSOLIDADO.xlsx' con las celdas de origen
# en los archivos individuales '[CODIGO].xlsx'.
# Clave: Letra de la columna en 'CONSOLIDADO.xlsx'
# Valor: Referencia de la celda en los archivos '[CODIGO].xlsx' (ej. 'N21')
COLUMNAS_DESTINO = {
    'F': 'N21',
    'G': 'N25',
    'I': 'N49',
    'K': 'N153',
    'M': 'N161'
}

# Lista de nombres de las hojas del archivo 'CONSOLIDADO.xlsx' que el script debe procesar.
HOJAS_OBJETIVO = ['EDU', 'HOSP', 'EMPRESA']

# La fila a partir de la cual el script comenzará a buscar códigos de datos.
# Las filas anteriores a esta se consideran encabezados.
FILA_INICIO = 6

# --- Funciones Auxiliares ---

def cargar_valor(wb, celda):
    """
    Intenta cargar el valor de una celda específica del libro de trabajo activo.

    Esta función es una envoltura segura para acceder a los valores de las celdas,
    evitando errores si la celda no existe o no tiene un valor.

    :param wb: El objeto Workbook de openpyxl del cual se extraerá el valor.
    :type wb: openpyxl.workbook.workbook.Workbook
    :param celda: La referencia de la celda (ej. 'A1', 'N21') cuyo valor se desea obtener.
    :type celda: str
    :returns: El valor de la celda si existe, de lo contrario, None.
    :rtype: any or None
    """
    try:
        return wb.active[celda].value
    except Exception:
        # Se captura cualquier excepción y se devuelve None para evitar que el script se detenga.
        return None

def procesar_hoja(ws, ruta_base):
    """
    Procesa una hoja específica del archivo 'CONSOLIDADO.xlsx'.

    Itera a través de las filas de la hoja, extrayendo un código de la
    COL_CODIGO (Columna D). Utiliza este código para buscar y leer un
    archivo Excel individual. Los valores de celdas predefinidas de este
    archivo individual se copian a las columnas de destino en la hoja actual.

    Los mensajes de estado se imprimen en la consola.

    :param ws: El objeto Worksheet de openpyxl de la hoja a procesar.
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param ruta_base: La ruta base donde se espera encontrar los archivos
    individuales '[CODIGO].xlsx'.
    :type ruta_base: str
    """
    fila = FILA_INICIO
    # Bucle infinito que se rompe cuando no se encuentra un código en la fila actual
    while True:
        # Obtener el valor de la celda en la columna del código para la fila actual
        codigo = ws.cell(row=fila, column=COL_CODIGO).value
        if not codigo:
            # Si la celda de código está vacía, se asume que no hay más datos y se sale del bucle
            break

        # Limpiar el código (convertir a string y eliminar espacios en blanco)
        codigo = str(codigo).strip()
        # Construir el nombre del archivo individual
        nombre_archivo = f"{codigo}.xlsx"
        # Construir la ruta completa al archivo individual
        ruta_archivo = os.path.join(ruta_base, nombre_archivo)

        # Verificar si el archivo individual existe
        if os.path.isfile(ruta_archivo):
            try:
                # Cargar el libro de trabajo del archivo individual en modo solo datos
                # (ignora estilos, fórmulas, etc. para una carga más rápida)
                wb_codigo = openpyxl.load_workbook(ruta_archivo, data_only=True)
                # Iterar sobre las columnas de destino y sus celdas de origen correspondientes
                for col_letra, celda in COLUMNAS_DESTINO.items():
                    # Cargar el valor de la celda de origen
                    valor = cargar_valor(wb_codigo, celda)
                    # Convertir la letra de la columna a su índice numérico (ej. 'F' -> 6)
                    col_index = openpyxl.utils.column_index_from_string(col_letra)
                    # Escribir el valor en la celda de destino en la hoja actual
                    ws.cell(row=fila, column=col_index, value=valor)
                print(f"  • Fila {fila:>4} | Código: {codigo:<10} | Estado: Copiado")
            except Exception as e:
                # Capturar y reportar cualquier error durante la lectura del archivo individual
                print(f"  • Fila {fila:>4} | Código: {codigo:<10} | Estado: Error al leer -> {e}")
        else:
            # Reportar si el archivo individual no fue encontrado
            print(f"  • Fila {fila:>4} | Código: {codigo:<10} | Estado: Archivo no encontrado")

        fila += 1 # Avanzar a la siguiente fila

# --- Función Principal ---

def main():
    """
    Función principal que orquesta el proceso de autocompletado y consolidación.

    Verifica la existencia del archivo 'CONSOLIDADO.xlsx', lo carga, itera
    sobre las hojas objetivo definidas, llama a 'procesar_hoja' para cada una,
    y finalmente guarda el libro de trabajo modificado bajo un nuevo nombre.

    NOTA IMPORTANTE: Esta función intenta modificar el archivo 'CONSOLIDADO.xlsx'
    directamente en memoria al cargarlo con `openpyxl.load_workbook(nombre_archivo)`.
    Para archivos Excel muy grandes (como el que causó el MemoryError en el pasado),
    esta aproximación es ineficiente y puede llevar a un 'MemoryError'.
    Las versiones más recientes del script utilizan un enfoque diferente
    (lectura en modo solo lectura y creación de un nuevo archivo de salida)
    para manejar archivos grandes de manera más robusta.
    """
    nombre_archivo = 'CONSOLIDADO.xlsx'
    archivo_salida = 'CONSOLIDADO_COMPLETADO.xlsx'

    print("Iniciando lectura y carga de datos...")

    # --- LÍNEAS PARA DEPURACIÓN ---
    """La lectura del archivo 'CONSOLIDADO.xlsx' puede tardar más de lo esperado 
    debido a loa complejidad del archivo, por eso se añadieros estas lineas de Debug, 
    para asegurarme que el proceso continuaba."""
    print(f"DEBUG: Directorio de trabajo actual: {os.getcwd()}")
    ruta_absoluta_consolidado = os.path.join(os.getcwd(), nombre_archivo)
    print(f"DEBUG: Intentando abrir: {ruta_absoluta_consolidado}")
    # ------

    # Verificar si el archivo 'CONSOLIDADO.xlsx' existe en el directorio actual
    if not os.path.exists(nombre_archivo):
        print(f"El archivo '{nombre_archivo}' no fue encontrado en el directorio de trabajo actual.")
        print("Asegúrate de que esté en la misma carpeta que el script/ejecutable.")
        input("Presiona ENTER para salir...") # Mantener la consola abierta hasta que el usuario presione Enter
        return # Terminar la ejecución

    # Intentar cargar el archivo 'CONSOLIDADO.xlsx'
    try:
        # Cargar el libro de trabajo completo.
        # Para archivos muy grandes, esta operación puede consumir mucha memoria.
        wb = openpyxl.load_workbook(nombre_archivo)
    except Exception as e:
        # Capturar y reportar errores si el archivo no se puede abrir
        print(f"No se pudo abrir el archivo Excel '{nombre_archivo}': {e}")
        print("Esto podría deberse a un archivo corrupto o problemas de permisos.")
        input("Presiona ENTER para salir...")
        return # Terminar la ejecución


    # Iterar sobre las hojas objetivo definidas en la constante HOJAS_OBJETIVO
    for hoja in HOJAS_OBJETIVO:
        # Verificar si la hoja existe en el libro de trabajo cargado
        if hoja not in wb.sheetnames:
            print(f"La hoja '{hoja}' no existe en el archivo. Saltando esta hoja.")
            continue # Pasar a la siguiente hoja si no se encuentra

        print(f"\nProcesando hoja: {hoja}")
        # Obtener el objeto de la hoja de trabajo actual
        ws = wb[hoja]
        # Llamar a la función para procesar la hoja
        procesar_hoja(ws, os.getcwd())
        print(f"Fin de hoja '{hoja}'")

    # Guardar el libro de trabajo modificado con un nuevo nombre
    try:
        wb.save(archivo_salida)
        print(f"\nArchivo guardado como '{archivo_salida}'")
    except Exception as e:
        print(f"Error al guardar el archivo '{archivo_salida}': {e}")
        print("Asegúrate de que el archivo no esté abierto en otra aplicación.")
    finally:
        # Este input se asegura de que la consola no se cierre inmediatamente
        # después de completar o fallar en el guardado.
        input("Presiona ENTER para salir...")

# Punto de entrada del script
if __name__ == '__main__':
    main()
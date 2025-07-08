"""
Script para autocompletar formulas en un archivo Excel maestro.

Este script lee un archivo Excel llamado 'CONSOLIDADO.xlsx', busca códigos
en la Columna D de hojas específicas ('EDU', 'HOSP', 'EMPRESA'), y utiliza
esos códigos para generar fórmulas de enlace a archivos Excel externos.

Al finalizar, el script guarda el archivo modificado con un nuevo nombre,
'CONSOLIDADO_COMPLETADO.xlsx'.

NOTA IMPORTANTE:

- Ahora inserta fórmulas con la **ruta absoluta completa fija** de los archivos
de origen (`C:\presupuestos`). Esto asegura que Excel pueda encontrar
los archivos si están en esa ubicación exacta.
- Asume que la hoja donde se encuentran los datos en los archivos
'[CODIGO].xlsx' SIEMPRE tiene el nombre especificado en
`NOMBRE_HOJA_ORIGEN_FIJO`.
"""

import openpyxl
import os

# --- Constantes de Configuración ---

# Define la columna donde se espera encontrar el código en las hojas objetivo (D = 4)
COL_CODIGO = 4  # Columna D

# Mapea las columnas de destino en 'CONSOLIDADO.xlsx' con las celdas de origen
# en los archivos individuales '[CODIGO].xlsx'.
# Clave: Letra de la columna en 'CONSOLIDADO.xlsx'
# Valor: Referencia de la celda en los archivos '[CODIGO].xlsx' (ej. '$N$21')
# Se usan signos '$' para hacer las referencias de celda absolutas.
COLUMNAS_DESTINO = {
    "F": "$N$21",
    "G": "$N$25",
    "I": "$N$49",
    "K": "$N$153",
    "M": "$N$161",
}

# Lista de nombres de las hojas del archivo 'CONSOLIDADO.xlsx' que el script debe procesar.
HOJAS_OBJETIVO = ["EDU", "HOSP", "EMPRESA"]

# La fila a partir de la cual el script comenzará a buscar códigos de datos.
# Las filas anteriores a esta se consideran encabezados.
FILA_INICIO = 6

NOMBRE_HOJA_ORIGEN_FIJO = "Hoja1" # Nombre de la hoja en los archivos '[CODIGO].xlsx' donde se encuentran los datos.

RUTA_ARCHIVOS_ORIGEN = r"C:\presupuestos" # Ruta absoluta fija donde se encuentran los archivos '[CODIGO].xlsx'

# --- Funciones Auxiliares (cargar_valor ya no es estrictamente necesaria aquí, pero se mantiene) ---


def cargar_valor(wb, celda):
    """
    Intenta cargar el valor de una celda específica del libro de trabajo activo.
    :param wb: El objeto Workbook de openpyxl del cual se extraerá el valor.
    :type wb: openpyxl.workbook.workbook.Workbook
    :param celda: La referencia de la celda (ej. 'A1', 'N21') cuyo valor se desea obtener.
    :type celda: str
    :returns: El valor de la celda si existe, de lo contrario, None.
    :rtype: any or None
    """
    try:
        return wb.active[celda].value
    except Exception:
        return None


def procesar_hoja(ws):
    """
    Procesa una hoja específica del archivo 'CONSOLIDADO.xlsx' insertando fórmulas
    de enlace externo a los archivos '[CODIGO].xlsx'.

    Itera a través de las filas de la hoja, extrayendo un código de la
    COL_CODIGO (Columna D). Utiliza este código para construir una
    fórmula de enlace externo con la ruta ABSOLUTA FIJA.
    Los mensajes de estado se imprimen en la consola.

    :param ws: El objeto Worksheet de openpyxl de la hoja a procesar.
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    """
    fila = FILA_INICIO
    # Bucle infinito que se rompe cuando no se encuentra un código en la fila actual
    while True:
        # Obtener el valor de la celda en la columna del código para la fila actual
        codigo = ws.cell(row=fila, column=COL_CODIGO).value
        if not codigo:
            # Si la celda de código está vacía, se asume que no hay más datos y se sale del bucle
            break

        # Limpiar el código (convertir a string y eliminar espacios en blanco)
        codigo = str(codigo).strip()
        # Construir el nombre del archivo individual
        nombre_archivo_codigo = f"{codigo}.xlsx"

        # OBTENER LA RUTA ABSOLUTA COMPLETA AL ARCHIVO DE ORIGEN USANDO LA CONSTANTE FIJA
        ruta_absoluta_archivo_codigo = os.path.abspath(
            os.path.join(RUTA_ARCHIVOS_ORIGEN, nombre_archivo_codigo)
        )

        # Extraer el directorio y el nombre del archivo base para la fórmula de Excel
        directorio_origen = os.path.dirname(ruta_absoluta_archivo_codigo)
        nombre_archivo_base = os.path.basename(ruta_absoluta_archivo_codigo)

        # Asegurar que el directorio termine con un separador de ruta para la sintaxis de Excel.
        path_para_formula = directorio_origen
        if directorio_origen and not directorio_origen.endswith(os.sep):
            path_para_formula += os.sep

        # Opcional: Verificar si el archivo existe para dar un mensaje informativo
        archivo_existe = os.path.isfile(ruta_absoluta_archivo_codigo)

        # Iterar sobre las columnas de destino y sus celdas de origen correspondientes
        for col_letra, celda_origen_ref in COLUMNAS_DESTINO.items():
            # Construir la fórmula de enlace externo con la RUTA ABSOLUTA FIJA
            # Formato: ='C:\Ruta\Al\Directorio\[NombreArchivo.xlsx]NombreHoja'!$Celda
            formula_enlace = f"='{path_para_formula}[{nombre_archivo_base}]{NOMBRE_HOJA_ORIGEN_FIJO}'!{celda_origen_ref}"

            # Convertir la letra de la columna a su índice numérico (ej. 'F' -> 6)
            col_index = openpyxl.utils.column_index_from_string(col_letra)
            # Escribir la FÓRMULA en la celda de destino
            ws.cell(row=fila, column=col_index, value=formula_enlace)

        if archivo_existe:
            print(
                f"  • Fila {fila:>4} | Código: {codigo:<10} | Estado: Fórmula ABSOLUTA añadida"
            )
        else:
            # Advertencia si el archivo no existe, aunque la fórmula se añade
            print(
                f"  • Fila {fila:>4} | Código: {codigo:<10} | Estado: Archivo de origen no encontrado, se añadió fórmula ABSOLUTA."
            )

        fila += 1  # Avanzar a la siguiente fila


# --- Función Principal ---


def main():
    """
    Función principal que orquesta el proceso de autocompletado y consolidación.

    Verifica la existencia del archivo 'CONSOLIDADO.xlsx', lo carga, itera
    sobre las hojas objetivo definidas, llama a 'procesar_hoja' para cada una,
    y finalmente guarda el libro de trabajo modificado bajo un nuevo nombre.

    NOTA IMPORTANTE: Esta función intenta modificar el archivo 'CONSOLIDADO.xlsx'
    directamente en memoria al cargarlo con `openpyxl.load_workbook(nombre_archivo)`.
    Para archivos Excel muy grandes, esta aproximación es ineficiente y puede
    llevar a un 'MemoryError'. Se recomienda la versión del script que crea un
    nuevo archivo de salida con modo `read_only=True` para archivos grandes.
    """
    nombre_archivo = "CONSOLIDADO.xlsx"
    archivo_salida = "CONSOLIDADO_COMPLETADO.xlsx"

    print("Iniciando lectura y carga de datos...")

    # --- LÍNEAS PARA DEPURACIÓN ---
    print(f"DEBUG: Directorio de trabajo actual: {os.getcwd()}")
    ruta_absoluta_consolidado = os.path.join(os.getcwd(), nombre_archivo)
    print(f"DEBUG: Intentando abrir: {ruta_absoluta_consolidado}")
    # ------

    # Verificar si el archivo 'CONSOLIDADO.xlsx' existe en el directorio actual
    if not os.path.exists(nombre_archivo):
        print(
            f"El archivo '{nombre_archivo}' no fue encontrado en el directorio de trabajo actual."
        )
        print("Asegúrate de que esté en la misma carpeta que el script/ejecutable.")
        input(
            "Presiona ENTER para salir..."
        )  # Mantener la consola abierta hasta que el usuario presione Enter
        return  # Terminar la ejecución

    # Intentar cargar el archivo 'CONSOLIDADO.xlsx'
    try:
        # Cargar el libro de trabajo completo.
        # NOTA: Para archivos muy grandes, esta operación puede consumir mucha memoria.
        wb = openpyxl.load_workbook(nombre_archivo)
    except Exception as e:
        # Capturar y reportar errores si el archivo no se puede abrir
        print(f"No se pudo abrir el archivo Excel '{nombre_archivo}': {e}")
        print("Esto podría deberse a un archivo corrupto o problemas de permisos.")
        input("Presiona ENTER para salir...")
        return  # Terminar la ejecución

    # Iterar sobre las hojas objetivo definidas en la constante HOJAS_OBJETIVO
    for hoja in HOJAS_OBJETIVO:
        # Verificar si la hoja existe en el libro de trabajo cargado
        if hoja not in wb.sheetnames:
            print(f"La hoja '{hoja}' no existe en el archivo. Saltando esta hoja.")
            continue  # Pasar a la siguiente hoja si no se encuentra

        print(f"\nProcesando hoja: {hoja}")
        # Obtener el objeto de la hoja de trabajo actual
        ws = wb[hoja]
        # Llamar a la función para procesar la hoja. Ya no necesita el argumento de ruta base.
        procesar_hoja(ws)
        print(f"Fin de hoja '{hoja}'")

    # Guardar el libro de trabajo modificado con un nuevo nombre
    try:
        wb.save(archivo_salida)
        print(f"\nArchivo guardado como '{archivo_salida}'")
    except Exception as e:
        print(f"Error al guardar el archivo '{archivo_salida}': {e}")
        print("Asegúrate de que el archivo no esté abierto en otra aplicación.")
    finally:
        # Este input se asegura de que la consola no se cierre inmediatamente
        # después de completar o fallar en el guardado.
        input("Presiona ENTER para salir...")


# Punto de entrada del script
if __name__ == "__main__":
    main()

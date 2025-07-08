"""
Script de Autocompletado y Consolidación de Datos Excel (Versión Optimizada para Memoria)

Este script automatiza el proceso de consolidar información de múltiples archivos Excel
individuales en un archivo maestro 'CONSOLIDADO.xlsx'. Lee el archivo maestro,
extrae códigos de la Columna D de hojas específicas ('EDU', 'HOSP', 'EMPRESA'),
y utiliza esos códigos para buscar archivos Excel individuales con el formato
'[CODIGO].xlsx'. Los datos de celdas predefinidas en estos archivos individuales
se copian luego a columnas específicas (F, G, I, K, M) en una nueva versión del
archivo consolidado.

Para manejar eficientemente archivos 'CONSOLIDADO.xlsx' de gran tamaño y evitar
problemas de 'MemoryError', el script opera de la siguiente manera:
1. Lee el archivo 'CONSOLIDADO.xlsx' en modo de solo lectura (streaming).
2. Procesa los datos y las modificaciones en memoria, fila por fila.
3. Escribe los datos (originales y modificados) en un *nuevo* archivo Excel de salida.

El resultado es un archivo `CONSOLIDADO_COMPLETADO.xlsx` que contiene todos los
datos originales y los nuevos, sin haber modificado el archivo maestro original
directamente en un modo que consuma mucha memoria.
"""

import openpyxl
import os
import sys
import traceback

# --- Constantes de Configuración ---

# Define la columna donde se espera encontrar el código en las hojas objetivo (D = 4).
# Nota: openpyxl usa indexación basada en 1 para las columnas.
COL_CODIGO = 4

# Mapea las columnas de destino en el archivo de salida con las celdas de origen
# correspondientes en los archivos individuales '[CODIGO].xlsx'.
# Clave (str): Letra de la columna en 'CONSOLIDADO_COMPLETADO.xlsx' (ej. 'F')
# Valor (str): Referencia de la celda en los archivos '[CODIGO].xlsx' (ej. 'N21')
COLUMNAS_DESTINO = {
    'F': 'N21',
    'G': 'N25',
    'I': 'N49',
    'K': 'N153',
    'M': 'N161'
}

# Diccionario auxiliar que convierte las letras de las columnas de destino a sus índices numéricos
# (para facilitar el acceso a las listas de filas de Python que son 0-indexed)
COL_INDICES_DESTINO = {
    openpyxl.utils.column_index_from_string(col_letra): celda
    for col_letra, celda in COLUMNAS_DESTINO.items()
}

# Lista de nombres de las hojas del archivo 'CONSOLIDADO.xlsx' que el script debe procesar
# para buscar y rellenar datos.
HOJAS_OBJETIVO = ['EDU', 'HOSP', 'EMPRESA']

# La fila a partir de la cual el script comenzará a buscar códigos de datos reales.
# Las filas anteriores a esta se consideran encabezados o metadatos y se copian directamente.
FILA_INICIO_DATOS = 6 # Fila 6 (1-indexed)

# --- Funciones Auxiliares ---

def cargar_valor_desde_origen(wb_origen, celda_ref):
    """
    Carga el valor de una celda específica del libro de trabajo de origen.

    Esta función es una envoltura segura para acceder a los valores de las celdas,
    capturando cualquier excepción y devolviendo None si la celda no existe o
    si ocurre algún otro problema al intentar acceder a su valor.

    :param wb_origen: El objeto Workbook de openpyxl del cual se extraerá el valor.
    Este Workbook debe estar abierto y activo.
    :type wb_origen: openpyxl.workbook.workbook.Workbook
    :param celda_ref: La referencia de la celda (ej. 'N21') cuyo valor se desea obtener.
    :type celda_ref: str
    :returns: El valor de la celda si existe, de lo contrario, None.
    :rtype: any or None
    """
    try:
        # Acceder a la hoja activa y a la celda por su referencia para obtener su valor.
        return wb_origen.active[celda_ref].value
    except Exception:
        # En caso de cualquier error (ej. celda no encontrada), se devuelve None.
        return None

def main():
    """
    Función principal que orquesta el proceso de autocompletado y consolidación.

    Gestiona la carga del archivo maestro 'CONSOLIDADO.xlsx' en modo de solo lectura
    para optimizar el uso de memoria. Itera sobre las hojas objetivo, procesa las
    filas para extraer y copiar datos de archivos externos, y finalmente guarda
    todos los datos (originales y modificados) en un nuevo archivo
    'CONSOLIDADO_COMPLETADO.xlsx'.

    También maneja la copia de hojas no objetivo del archivo original al nuevo
    archivo de salida y proporciona mensajes de depuración y error en la consola.
    """
    # Nombres de los archivos de entrada y salida
    nombre_archivo_consolidado = 'CONSOLIDADO.xlsx'
    nombre_archivo_salida = 'CONSOLIDADO_COMPLETADO.xlsx'
    # Obtener el directorio de trabajo actual donde se espera que estén todos los archivos.
    ruta_base = os.getcwd()

    print("Iniciando lectura y carga de datos...")
    print(f"DEBUG: Directorio de trabajo actual: {ruta_base}")
    # Construir la ruta absoluta al archivo consolidado.
    ruta_absoluta_consolidado = os.path.join(ruta_base, nombre_archivo_consolidado)
    print(f"DEBUG: Intentando procesar: {ruta_absoluta_consolidado}")

    # Verificar la existencia del archivo 'CONSOLIDADO.xlsx'.
    if not os.path.exists(ruta_absoluta_consolidado):
        print(f"El archivo '{nombre_archivo_consolidado}' no fue encontrado en el directorio de trabajo actual.")
        print("Por favor, asegúrate de que esté en la misma carpeta que el ejecutable o el script.")
        input("Presiona ENTER para salir...") # Mantener la consola abierta para que el usuario pueda leer el error.
        return # Salir del script.

    # Crear un nuevo libro de trabajo para almacenar los datos procesados.
    # Este será el archivo 'CONSOLIDADO_COMPLETADO.xlsx'.
    wb_salida = openpyxl.Workbook()
    # Eliminar la hoja por defecto 'Sheet' que openpyxl crea si no hay otras hojas aún.
    if 'Sheet' in wb_salida.sheetnames and len(wb_salida.sheetnames) == 1:
        wb_salida.remove(wb_salida['Sheet'])

    # Inicializar la variable del libro de trabajo de lectura a None.
    # Esto es útil para asegurar que wb_consolidado_lectura siempre esté definida
    # y pueda ser cerrada en el bloque 'finally' incluso si la carga falla.
    wb_consolidado_lectura = None
    try:
        # --- CARGA DEL ARCHIVO CONSOLIDADO EN MODO DE SOLO LECTURA ---
        # Este es el paso clave para la optimización de memoria.
        # openpyxl leerá el archivo de forma eficiente sin cargar todo en RAM.
        print(f"DEBUG: Cargando '{nombre_archivo_consolidado}' en modo solo lectura...")
        wb_consolidado_lectura = openpyxl.load_workbook(ruta_absoluta_consolidado, read_only=True)
        print("DEBUG: Archivo consolidado cargado en modo solo lectura con éxito.")

        # Iterar sobre cada hoja definida en HOJAS_OBJETIVO para su procesamiento.
        for nombre_hoja_objetivo in HOJAS_OBJETIVO:
            # Verificar si la hoja objetivo existe en el archivo consolidado.
            if nombre_hoja_objetivo not in wb_consolidado_lectura.sheetnames:
                print(f"La hoja '{nombre_hoja_objetivo}' no existe en '{nombre_archivo_consolidado}'. Saltando.")
                continue # Saltar a la siguiente hoja si no se encuentra.

            print(f"\nProcesando hoja: {nombre_hoja_objetivo}")
            # Obtener el objeto de la hoja de lectura en modo solo lectura.
            ws_lectura = wb_consolidado_lectura[nombre_hoja_objetivo]

            # Crear una nueva hoja en el libro de trabajo de salida o seleccionarla si ya existe.
            if nombre_hoja_objetivo in wb_salida.sheetnames:
                ws_salida = wb_salida[nombre_hoja_objetivo]
            else:
                ws_salida = wb_salida.create_sheet(title=nombre_hoja_objetivo)

            # Iterar sobre las filas de la hoja original desde la primera fila (min_row=1).
            # `values_only=False` es importante para acceder a los objetos Cell y no solo a sus valores,
            # lo que permite acceder a `cell.value` y `cell.column`.
            for r_idx, row_data in enumerate(ws_lectura.iter_rows(min_row=1, values_only=False)):
                # Convertir la fila de objetos Cell a una lista de sus valores.
                # Esta lista representará la fila que se escribirá en el nuevo archivo.
                current_row_values = [cell.value for cell in row_data]
                
                # Las filas antes de FILA_INICIO_DATOS se consideran encabezados.
                # Se copian directamente al nuevo archivo sin procesamiento de códigos.
                if r_idx + 1 < FILA_INICIO_DATOS: # r_idx es 0-indexed, fila es 1-indexed.
                    ws_salida.append(current_row_values) # Escribir la fila de encabezado directamente.
                    continue # Pasar a la siguiente fila.

                # --- Lógica de procesamiento de datos para filas a partir de FILA_INICIO_DATOS ---
                # Convertir el índice de fila basado en 0 a un número de fila basado en 1 para los logs.
                fila_actual_num = r_idx + 1
                
                # Obtener el objeto Cell de la columna de código (COL_CODIGO) de la fila actual.
                # Se resta 1 porque `row_data` es una lista 0-indexed.
                codigo_cell = row_data[COL_CODIGO - 1]
                # Extraer el valor del código, convertirlo a string y eliminar espacios en blanco.
                codigo = str(codigo_cell.value).strip() if codigo_cell.value else ""
                
                # Si no se encuentra un código en la celda, la fila se copia tal cual sin buscar un archivo externo.
                if not codigo:
                    ws_salida.append(current_row_values)
                    continue # Pasar a la siguiente fila.

                # Construir la ruta completa al archivo individual '[CODIGO].xlsx'.
                ruta_archivo_origen = os.path.join(ruta_base, f"{codigo}.xlsx")
                
                # Verificar si el archivo individual existe.
                if os.path.isfile(ruta_archivo_origen):
                    try:
                        # Cargar el libro de trabajo del archivo individual en modo solo datos.
                        wb_codigo = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
                        # Iterar sobre las columnas de destino definidas en COL_INDICES_DESTINO.
                        for col_index, celda_ref_origen in COL_INDICES_DESTINO.items():
                            # Cargar el valor de la celda de origen del archivo individual.
                            valor = cargar_valor_desde_origen(wb_codigo, celda_ref_origen)
                            # Asegurarse de que la lista 'current_row_values' tenga suficientes elementos
                            # para la columna de destino. Si es más corta, se rellena con None.
                            while len(current_row_values) < col_index:
                                current_row_values.append(None)
                            # Actualizar el valor en la posición correcta de la lista de la fila.
                            # Se resta 1 porque 'col_index' es 1-indexed y la lista es 0-indexed.
                            current_row_values[col_index - 1] = valor
                        print(f"  • Fila {fila_actual_num:>4} | Código: {codigo:<10} | Estado: Copiado")
                    except Exception as e:
                        # Reportar errores específicos al leer el archivo individual.
                        print(f"  • Fila {fila_actual_num:>4} | Código: {codigo:<10} | Estado: Error al leer -> {e}")
                else:
                    # Reportar si el archivo individual no fue encontrado.
                    print(f"  • Fila {fila_actual_num:>4} | Código: {codigo:<10} | Estado: Archivo no encontrado")
                
                # Añadir la fila (ya sea modificada o copiada tal cual) a la hoja de salida.
                ws_salida.append(current_row_values)

            print(f"Fin de hoja '{nombre_hoja_objetivo}'")
        
        # --- Copia de Hojas No Procesadas ---
        # Después de procesar las hojas objetivo, se asegura que cualquier otra hoja
        # del archivo original que no haya sido tocada también se copie al nuevo libro.
        for existing_sheet_name in wb_consolidado_lectura.sheetnames:
            # Si la hoja no estaba en la lista de HOJAS_OBJETIVO y aún no ha sido creada en la salida.
            if existing_sheet_name not in HOJAS_OBJETIVO and existing_sheet_name not in wb_salida.sheetnames:
                print(f"DEBUG: Copiando hoja no procesada: {existing_sheet_name}")
                # Obtener la hoja original del libro de lectura.
                ws_original = wb_consolidado_lectura[existing_sheet_name]
                # Crear una nueva hoja con el mismo nombre en el libro de salida.
                ws_nueva = wb_salida.create_sheet(title=existing_sheet_name)
                # Copiar todas las filas de la hoja original a la nueva hoja de salida.
                for row in ws_original.iter_rows(values_only=True):
                    ws_nueva.append(row)

    except Exception as e:
        # --- Manejo de ERRORES CRÍTICOS DURANTE EL PROCESAMIENTO ---
        print(f"ERROR CRÍTICO durante el procesamiento: {e}")
        print("\n--- INICIO DE TRACEBACK DETALLADO ---")
        # Imprimir el traceback completo para depuración.
        traceback.print_exc()
        print("--- FIN DE TRACEBACK DETALLADO ---\n")
        print("Esto podría deberse a un archivo Excel corrupto, problemas de permisos, o un problema inesperado.")
        input("Presiona ENTER para salir...")
        sys.exit(1) # Salir del script con un código de error.
    finally:
        # --- Cierre de Archivo de Lectura ---
        # Este bloque 'finally' asegura que el libro de trabajo de lectura siempre
        # se cierre, incluso si ocurre una excepción, liberando recursos.
        if wb_consolidado_lectura is not None:
            wb_consolidado_lectura.close()


    # --- Guarda el Nuevo Libro de Trabajo ---
    ruta_absoluta_salida = os.path.join(ruta_base, nombre_archivo_salida)
    try:
        # Intentar guardar el libro de trabajo de salida.
        wb_salida.save(ruta_absoluta_salida)
        print(f"\nArchivo modificado guardado como '{os.path.basename(ruta_absoluta_salida)}'")
    except Exception as e:
        # Capturar y reportar errores durante el proceso de guardado.
        print(f"Error al guardar el archivo '{os.path.basename(ruta_absoluta_salida)}': {e}")
        print("Asegúrate de que el archivo no esté abierto en otra aplicación o que tengas permisos de escritura.")
        input("Presiona ENTER para salir...")
        sys.exit(1) # Salir del script con un código de error.

    print("\n El proceso ha terminado correctamente.")
    input("Presiona ENTER para salir...") # Pausar la consola al finalizar.

# Punto de entrada del script. Asegura que `main()` se ejecute solo cuando el script
# es ejecutado directamente y no cuando es importado como un módulo.
if __name__ == '__main__':
    main()